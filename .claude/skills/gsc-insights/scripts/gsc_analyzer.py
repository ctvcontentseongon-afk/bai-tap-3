"""
SEONGON GSC Insights Analyzer — Phân tích Google Search Console chuyên sâu
Output 100% tiếng Việt | Triết lý: WHAT → WHY → HOW
"""

import argparse
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Đường dẫn
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent.parent
TESTS_DIR = BASE_DIR / "tests"
OUTPUTS_DIR = BASE_DIR.parent.parent.parent / "outputs"

# ---------------------------------------------------------------------------
# Bảng chuẩn tỷ lệ nhấp (CTR benchmark) theo vị trí
# ---------------------------------------------------------------------------
CTR_BENCHMARK = {
    1: 0.285, 2: 0.157, 3: 0.110, 4: 0.080, 5: 0.072,
    6: 0.051, 7: 0.047, 8: 0.038, 9: 0.035, 10: 0.029,
    11: 0.018, 12: 0.018, 13: 0.018, 14: 0.018, 15: 0.018,
}

def lay_chuan_ctr(vi_tri: float) -> float:
    vi_tri_lam_tron = max(1, min(15, round(vi_tri)))
    return CTR_BENCHMARK.get(vi_tri_lam_tron, 0.010)

# ---------------------------------------------------------------------------
# Đọc dữ liệu mock
# ---------------------------------------------------------------------------
def doc_du_lieu_mock(ten_file: str) -> pd.DataFrame:
    duong_dan = TESTS_DIR / ten_file
    if not duong_dan.exists():
        raise FileNotFoundError(f"Không tìm thấy file mock: {duong_dan}")
    df = pd.read_csv(duong_dan)
    df.columns = [c.strip() for c in df.columns]
    df = df.fillna("")
    df["clicks"] = pd.to_numeric(df["clicks"], errors="coerce").fillna(0).astype(int)
    df["impressions"] = pd.to_numeric(df["impressions"], errors="coerce").fillna(0).astype(int)
    df["ctr"] = pd.to_numeric(df["ctr"], errors="coerce").fillna(0.0)
    df["position"] = pd.to_numeric(df["position"], errors="coerce").fillna(0.0)
    return df

# ---------------------------------------------------------------------------
# 1. CƠ HỘI TỐI ƯU CTR NHANH
# ---------------------------------------------------------------------------
def tim_co_hoi_toi_uu(df: pd.DataFrame) -> pd.DataFrame:
    """Vị trí 5–15, lượt hiển thị ≥ 100, CTR thực tế < 80% chuẩn."""
    ket_qua = []
    for _, row in df.iterrows():
        vi_tri = row["position"]
        luot_hien_thi = row["impressions"]
        ctr_thuc_te = row["ctr"]

        if not (5.0 <= vi_tri <= 15.0 and luot_hien_thi >= 100):
            continue

        chuan = lay_chuan_ctr(vi_tri)
        if ctr_thuc_te >= chuan * 0.80:
            continue

        chenh_lech = chuan - ctr_thuc_te
        clicks_tiem_nang = round(luot_hien_thi * chenh_lech)
        diem_uu_tien = round(luot_hien_thi * chenh_lech * 100)

        nguyen_nhan = _chan_doan_ctr_thap(vi_tri, ctr_thuc_te, chuan, luot_hien_thi)
        hanh_dong = _de_xuat_tang_ctr(vi_tri, nguyen_nhan)

        ket_qua.append({
            "Từ khóa": row["query"],
            "Trang đích": row["page"],
            "Vị trí TB": round(vi_tri, 1),
            "Lượt hiển thị": luot_hien_thi,
            "CTR thực tế": f"{ctr_thuc_te:.1%}",
            "CTR chuẩn": f"{chuan:.1%}",
            "Clicks tiềm năng": clicks_tiem_nang,
            "Điểm ưu tiên": diem_uu_tien,
            "Nguyên nhân CTR thấp": nguyen_nhan,
            "Hành động đề xuất": hanh_dong,
        })

    df_ket_qua = pd.DataFrame(ket_qua)
    if not df_ket_qua.empty:
        df_ket_qua = df_ket_qua.sort_values("Điểm ưu tiên", ascending=False).reset_index(drop=True)
    return df_ket_qua


def _chan_doan_ctr_thap(vi_tri, ctr_thuc_te, chuan, luot_hien_thi):
    ti_le = ctr_thuc_te / chuan if chuan > 0 else 0
    if ti_le < 0.3:
        return "Title/meta description kém hấp dẫn hoặc không khớp search intent"
    elif vi_tri < 8 and ti_le < 0.5:
        return "Snippet chưa nổi bật, thiếu số liệu/year/power word trong title"
    elif luot_hien_thi > 1500 and ti_le < 0.5:
        return "Từ khóa cạnh tranh cao, SERP có nhiều featured snippet/ads che phần organic"
    elif vi_tri > 10:
        return "Vị trí thấp (trang 2), cần cải thiện để lên top 10 trước"
    else:
        return "Meta description thiếu CTA hoặc chưa chứa từ khóa chính"


def _de_xuat_tang_ctr(vi_tri, nguyen_nhan):
    if "title" in nguyen_nhan.lower() or "snippet" in nguyen_nhan.lower():
        return "Viết lại title tag: thêm con số, năm hiện tại, power word (Đầy đủ, Chi tiết, [2026])"
    elif "trang 2" in nguyen_nhan or vi_tri > 10:
        return "Ưu tiên tăng thứ hạng lên top 10 bằng cách bổ sung nội dung và backlink trước"
    elif "featured snippet" in nguyen_nhan or "ads" in nguyen_nhan:
        return "Tối ưu meta description 150–160 ký tự, thêm schema FAQ để chiếm rich result"
    else:
        return "A/B test 2 phiên bản title khác nhau trong Google Search Console"

# ---------------------------------------------------------------------------
# 2. TRÙNG LẶP NỘI DUNG
# ---------------------------------------------------------------------------
def tim_trung_lap_noi_dung(df: pd.DataFrame) -> pd.DataFrame:
    """Từ khóa có >1 URL trong top 20."""
    nhom = defaultdict(list)
    for _, row in df.iterrows():
        if row["position"] <= 20:
            nhom[row["query"]].append(row)

    ket_qua = []
    for query, danh_sach in nhom.items():
        if len(danh_sach) <= 1:
            continue

        danh_sach = sorted(danh_sach, key=lambda x: x["position"])
        trang_chinh = danh_sach[0]
        cac_trang_khac = danh_sach[1:]

        tong_impressions = sum(r["impressions"] for r in danh_sach)
        tong_clicks = sum(r["clicks"] for r in danh_sach)

        so_url = len(danh_sach)
        chenh_lech_vi_tri = danh_sach[-1]["position"] - danh_sach[0]["position"]
        muc_do, ly_do = _danh_gia_trung_lap(so_url, chenh_lech_vi_tri, tong_impressions)
        giai_phap = _de_xuat_xu_ly_trung_lap(
            trang_chinh["page"], [r["page"] for r in cac_trang_khac]
        )

        ket_qua.append({
            "Từ khóa": query,
            "Số URL cạnh tranh": so_url,
            "Trang chính (thứ hạng cao nhất)": trang_chinh["page"],
            "Vị trí trang chính": round(trang_chinh["position"], 1),
            "Các URL cạnh tranh": " | ".join(r["page"] for r in cac_trang_khac),
            "Tổng lượt hiển thị": tong_impressions,
            "Tổng lượt nhấp": tong_clicks,
            "Mức độ nghiêm trọng": muc_do,
            "Nguyên nhân gây trùng lặp": ly_do,
            "Giải pháp đề xuất": giai_phap,
        })

    df_ket_qua = pd.DataFrame(ket_qua)
    if not df_ket_qua.empty:
        df_ket_qua = df_ket_qua.sort_values("Tổng lượt hiển thị", ascending=False).reset_index(drop=True)
    return df_ket_qua


def _danh_gia_trung_lap(so_url, chenh_lech_vi_tri, tong_impressions):
    if so_url >= 3 or (so_url == 2 and chenh_lech_vi_tri < 5):
        return "Nghiêm trọng", "Nhiều trang cạnh tranh trực tiếp, Google không xác định được trang canonical"
    elif tong_impressions > 2000:
        return "Trung bình", "Từ khóa quan trọng bị phân tán lực, ảnh hưởng đến thứ hạng tổng thể"
    else:
        return "Nhẹ", "Overlap nội dung, cần theo dõi thêm trước khi xử lý"


def _de_xuat_xu_ly_trung_lap(trang_chinh, cac_trang_khac):
    return (
        f"Giữ {trang_chinh} làm canonical. "
        f"Thêm rel=canonical hoặc 301 redirect từ: {', '.join(cac_trang_khac[:2])}. "
        "Consolidate nội dung nếu 2 trang quá tương đồng."
    )

# ---------------------------------------------------------------------------
# 3. TỪ KHÓA TỤT HẠNG
# ---------------------------------------------------------------------------
def tim_tu_khoa_tut_hang(df_hien_tai: pd.DataFrame, df_truoc: pd.DataFrame) -> pd.DataFrame:
    """So sánh 2 kỳ, lọc từ khóa tụt > 2 bậc."""
    def tong_hop(df):
        return (
            df.groupby(["query", "page"])
            .agg(vi_tri=("position", "mean"), luot_hien_thi=("impressions", "sum"), luot_nhap=("clicks", "sum"))
            .reset_index()
        )

    hop_nhat = tong_hop(df_hien_tai).merge(
        tong_hop(df_truoc), on=["query", "page"], suffixes=("_ht", "_t")
    )

    ket_qua = []
    for _, row in hop_nhat.iterrows():
        chenh_lech = row["vi_tri_ht"] - row["vi_tri_t"]
        if chenh_lech <= 2.0:
            continue

        ket_qua.append({
            "Từ khóa": row["query"],
            "Trang đích": row["page"],
            "Vị trí kỳ trước": round(row["vi_tri_t"], 1),
            "Vị trí kỳ này": round(row["vi_tri_ht"], 1),
            "Tụt (số bậc)": round(chenh_lech, 1),
            "Lượt hiển thị kỳ này": int(row["luot_hien_thi_ht"]),
            "Thay đổi lượt nhấp": int(row["luot_nhap_ht"] - row["luot_nhap_t"]),
            "Mức độ nghiêm trọng": _muc_do_tut(chenh_lech),
            "Nguyên nhân có thể": _phan_tich_nguyen_nhan_tut_hang(
                row["vi_tri_t"], row["vi_tri_ht"], chenh_lech, row["luot_hien_thi_ht"]
            ),
            "Kế hoạch phục hồi": _de_xuat_phuc_hoi(chenh_lech, row["vi_tri_ht"]),
        })

    df_ket_qua = pd.DataFrame(ket_qua)
    if not df_ket_qua.empty:
        df_ket_qua = df_ket_qua.sort_values("Tụt (số bậc)", ascending=False).reset_index(drop=True)
    return df_ket_qua


def _muc_do_tut(chenh_lech):
    if chenh_lech >= 10:
        return "Khủng hoảng (≥10 bậc)"
    elif chenh_lech >= 5:
        return "Nghiêm trọng (5–9 bậc)"
    else:
        return "Cần chú ý (2–4 bậc)"


def _phan_tich_nguyen_nhan_tut_hang(vi_tri_cu, vi_tri_moi, chenh_lech, luot_hien_thi):
    if vi_tri_cu <= 5 and vi_tri_moi > 10:
        return "Có thể bị Google algorithm update ảnh hưởng hoặc đối thủ tăng tốc mạnh"
    elif chenh_lech >= 10:
        return "Nội dung lỗi thời, thiếu E-E-A-T, hoặc trang bị lỗi kỹ thuật (crawl error)"
    elif luot_hien_thi < 500:
        return "Từ khóa niche đang thay đổi search volume theo mùa"
    else:
        return "Cạnh tranh tăng: đối thủ cập nhật nội dung mới hoặc xây backlink chất lượng hơn"


def _de_xuat_phuc_hoi(chenh_lech, vi_tri_moi):
    if vi_tri_moi > 15:
        return "Audit nội dung: kiểm tra E-E-A-T, cập nhật thông tin mới nhất, bổ sung schema"
    elif chenh_lech >= 5:
        return "Phân tích top 3 đối thủ, cải thiện nội dung và xây thêm backlink trong 30 ngày"
    else:
        return "Theo dõi thêm 2 tuần, nếu tiếp tục tụt thì tiến hành content refresh"

# ---------------------------------------------------------------------------
# 4. TỪ KHÓA MỚI NỔI
# ---------------------------------------------------------------------------
def tim_tu_khoa_moi_noi(df_hien_tai: pd.DataFrame, df_truoc: pd.DataFrame) -> pd.DataFrame:
    """Từ khóa mới trong top 50, hoặc tăng ≥ 5 bậc."""
    def tong_hop(df):
        return (
            df.groupby(["query", "page"])
            .agg(vi_tri=("position", "mean"), luot_hien_thi=("impressions", "sum"), luot_nhap=("clicks", "sum"))
            .reset_index()
        )

    hien_tai_tong = tong_hop(df_hien_tai)
    truoc_tong = tong_hop(df_truoc)
    cap_truoc = set(zip(df_truoc["query"], df_truoc["page"]))

    ket_qua = []

    # Từ khóa mới hoàn toàn
    for _, row in hien_tai_tong.iterrows():
        if (row["query"], row["page"]) in cap_truoc or row["vi_tri"] > 50:
            continue
        ket_qua.append({
            "Từ khóa": row["query"],
            "Trang đích": row["page"],
            "Loại": "Từ khóa mới",
            "Vị trí kỳ trước": "—",
            "Vị trí kỳ này": round(row["vi_tri"], 1),
            "Tăng (số bậc)": "Mới",
            "Lượt hiển thị": int(row["luot_hien_thi"]),
            "Lượt nhấp": int(row["luot_nhap"]),
            "Tiềm năng": _danh_gia_tiem_nang(row["vi_tri"], row["luot_hien_thi"]),
            "Hành động khai thác": _de_xuat_khai_thac(row["vi_tri"], row["luot_hien_thi"]),
        })

    # Từ khóa tăng hạng mạnh
    hop_nhat = hien_tai_tong.merge(truoc_tong, on=["query", "page"], suffixes=("_ht", "_t"))
    for _, row in hop_nhat.iterrows():
        tang = row["vi_tri_t"] - row["vi_tri_ht"]
        if tang < 5:
            continue
        ket_qua.append({
            "Từ khóa": row["query"],
            "Trang đích": row["page"],
            "Loại": "Tăng hạng mạnh",
            "Vị trí kỳ trước": round(row["vi_tri_t"], 1),
            "Vị trí kỳ này": round(row["vi_tri_ht"], 1),
            "Tăng (số bậc)": round(tang, 1),
            "Lượt hiển thị": int(row["luot_hien_thi_ht"]),
            "Lượt nhấp": int(row["luot_nhap_ht"]),
            "Tiềm năng": _danh_gia_tiem_nang(row["vi_tri_ht"], row["luot_hien_thi_ht"]),
            "Hành động khai thác": _de_xuat_khai_thac(row["vi_tri_ht"], row["luot_hien_thi_ht"]),
        })

    df_ket_qua = pd.DataFrame(ket_qua)
    if not df_ket_qua.empty:
        df_ket_qua = df_ket_qua.sort_values("Lượt hiển thị", ascending=False).reset_index(drop=True)
    return df_ket_qua


def _danh_gia_tiem_nang(vi_tri, luot_hien_thi):
    if vi_tri <= 10 and luot_hien_thi > 1000:
        return "Cao — Khai thác ngay"
    elif vi_tri <= 20 and luot_hien_thi > 500:
        return "Trung bình — Đầu tư content"
    else:
        return "Theo dõi — Chờ thêm data"


def _de_xuat_khai_thac(vi_tri, luot_hien_thi):
    if vi_tri <= 10:
        return "Tối ưu CTR (title/meta), thêm internal link để giữ vị trí"
    elif vi_tri <= 20:
        return "Bổ sung nội dung chuyên sâu, build thêm 2–3 backlink chất lượng"
    else:
        return "Theo dõi 4 tuần, nếu ổn định thì đầu tư tối ưu toàn diện"

# ---------------------------------------------------------------------------
# 5. TRANG HIỆU QUẢ NHẤT
# ---------------------------------------------------------------------------
def phan_tich_trang_hieu_qua(df: pd.DataFrame) -> pd.DataFrame:
    tong_hop = (
        df.groupby("page")
        .agg(tong_clicks=("clicks", "sum"), tong_impressions=("impressions", "sum"),
             vi_tri_tb=("position", "mean"), so_tu_khoa=("query", "count"))
        .reset_index()
    )
    tong_hop["CTR thực tế"] = tong_hop.apply(
        lambda r: f"{r['tong_clicks'] / r['tong_impressions']:.1%}" if r["tong_impressions"] > 0 else "0.0%",
        axis=1,
    )
    tong_hop = tong_hop.sort_values("tong_clicks", ascending=False).head(20).reset_index(drop=True)
    tong_hop = tong_hop.rename(columns={
        "page": "Trang",
        "tong_clicks": "Tổng lượt nhấp",
        "tong_impressions": "Tổng lượt hiển thị",
        "vi_tri_tb": "Vị trí TB",
        "so_tu_khoa": "Số từ khóa",
    })
    tong_hop["Vị trí TB"] = tong_hop["Vị trí TB"].round(1)
    return tong_hop[["Trang", "Tổng lượt nhấp", "Tổng lượt hiển thị", "CTR thực tế", "Vị trí TB", "Số từ khóa"]]

# ---------------------------------------------------------------------------
# 6. TỪ KHÓA HIỆU QUẢ NHẤT
# ---------------------------------------------------------------------------
def phan_tich_tu_khoa_hieu_qua(df: pd.DataFrame) -> pd.DataFrame:
    tong_hop = (
        df.groupby("query")
        .agg(tong_clicks=("clicks", "sum"), tong_impressions=("impressions", "sum"),
             vi_tri_tb=("position", "mean"))
        .reset_index()
    )
    tong_hop["CTR thực tế"] = tong_hop.apply(
        lambda r: f"{r['tong_clicks'] / r['tong_impressions']:.1%}" if r["tong_impressions"] > 0 else "0.0%",
        axis=1,
    )
    tong_hop = tong_hop.sort_values("tong_clicks", ascending=False).head(20).reset_index(drop=True)
    tong_hop = tong_hop.rename(columns={
        "query": "Từ khóa",
        "tong_clicks": "Tổng lượt nhấp",
        "tong_impressions": "Tổng lượt hiển thị",
        "vi_tri_tb": "Vị trí TB",
    })
    tong_hop["Vị trí TB"] = tong_hop["Vị trí TB"].round(1)
    return tong_hop[["Từ khóa", "Tổng lượt nhấp", "Tổng lượt hiển thị", "CTR thực tế", "Vị trí TB"]]

# ---------------------------------------------------------------------------
# 7. TỔNG QUAN 2 KỲ
# ---------------------------------------------------------------------------
def tong_hop_toan_canh(df_hien_tai: pd.DataFrame, df_truoc: pd.DataFrame) -> dict:
    def kpi(df):
        tong_clicks = int(df["clicks"].sum())
        tong_impr = int(df["impressions"].sum())
        return {
            "tong_clicks": tong_clicks,
            "tong_impressions": tong_impr,
            "ctr_tb": tong_clicks / tong_impr if tong_impr > 0 else 0,
            "vi_tri_tb": float(df["position"].mean()),
            "so_tu_khoa": int(df["query"].nunique()),
        }

    ht = kpi(df_hien_tai)
    t = kpi(df_truoc)

    def thay_doi_pct(a, b):
        if b == 0:
            return "N/A"
        d = (a - b) / b
        return f"{'▲' if d > 0 else '▼'} {abs(d):.1%}"

    def thay_doi_abs(a, b, fmt=".1f"):
        d = a - b
        return f"{'▲' if d > 0 else '▼'} {abs(d):{fmt}}"

    return {
        "Tổng lượt nhấp": {"Kỳ này": ht["tong_clicks"], "Kỳ trước": t["tong_clicks"], "Thay đổi": thay_doi_pct(ht["tong_clicks"], t["tong_clicks"])},
        "Tổng lượt hiển thị": {"Kỳ này": ht["tong_impressions"], "Kỳ trước": t["tong_impressions"], "Thay đổi": thay_doi_pct(ht["tong_impressions"], t["tong_impressions"])},
        "CTR trung bình": {"Kỳ này": f"{ht['ctr_tb']:.2%}", "Kỳ trước": f"{t['ctr_tb']:.2%}", "Thay đổi": thay_doi_abs(ht["ctr_tb"] * 100, t["ctr_tb"] * 100, ".2f") + " điểm %"},
        "Vị trí trung bình": {"Kỳ này": round(ht["vi_tri_tb"], 1), "Kỳ trước": round(t["vi_tri_tb"], 1), "Thay đổi": thay_doi_abs(ht["vi_tri_tb"], t["vi_tri_tb"])},
        "Số từ khóa": {"Kỳ này": ht["so_tu_khoa"], "Kỳ trước": t["so_tu_khoa"], "Thay đổi": thay_doi_pct(ht["so_tu_khoa"], t["so_tu_khoa"])},
    }

# ---------------------------------------------------------------------------
# XUẤT EXCEL — 7 SHEET
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(fgColor="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
BORDER_THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
FILL_ZEBRA = PatternFill(fgColor="EBF3FB", fill_type="solid")
COL_WRAP = {
    "Nguyên nhân CTR thấp", "Hành động đề xuất", "Nguyên nhân có thể",
    "Kế hoạch phục hồi", "Nguyên nhân gây trùng lặp", "Giải pháp đề xuất",
    "Hành động khai thác", "Các URL cạnh tranh",
}


def _ghi_sheet(ws, df: pd.DataFrame):
    if df.empty:
        ws.append(["(Không có dữ liệu)"])
        return

    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_THIN
    ws.row_dimensions[1].height = 35
    ws.freeze_panes = "A2"

    for i, (_, row) in enumerate(df.iterrows(), start=2):
        ws.append(list(row))
        fill = FILL_ZEBRA if i % 2 == 0 else None
        for j, cell in enumerate(ws[i], start=1):
            col_name = df.columns[j - 1]
            cell.alignment = Alignment(vertical="top", wrap_text=(col_name in COL_WRAP))
            cell.border = BORDER_THIN
            if fill:
                cell.fill = fill

    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        if col_name in COL_WRAP:
            ws.column_dimensions[col_letter].width = 48
        elif "Trang" in col_name or "URL" in col_name:
            ws.column_dimensions[col_letter].width = 52
        elif "Từ khóa" in col_name:
            ws.column_dimensions[col_letter].width = 32
        else:
            max_len = max(len(str(cell.value or "")) for cell in ws[col_letter])
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 30)


def _ghi_sheet_tong_quan(ws, tong_quan: dict):
    ws.append(["Chỉ số", "Kỳ này", "Kỳ trước", "Thay đổi"])
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER_THIN
    ws.row_dimensions[1].height = 35
    ws.freeze_panes = "A2"

    for i, (chi_so, gia_tri) in enumerate(tong_quan.items(), start=2):
        ws.append([chi_so, gia_tri["Kỳ này"], gia_tri["Kỳ trước"], gia_tri["Thay đổi"]])
        fill = FILL_ZEBRA if i % 2 == 0 else None
        for cell in ws[i]:
            cell.alignment = Alignment(vertical="center")
            cell.border = BORDER_THIN
            if fill:
                cell.fill = fill

    for col, w in [("A", 26), ("B", 18), ("C", 18), ("D", 20)]:
        ws.column_dimensions[col].width = w


def xuat_excel(ten_mien, tong_quan, co_hoi, trung_lap, tut_hang, moi_noi,
               trang_hieu_qua, tu_khoa_hieu_qua, thu_muc_xuat, ngay):
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Tổng quan"
    _ghi_sheet_tong_quan(ws1, tong_quan)

    for ten, df in [
        ("Cơ hội tối ưu nhanh", co_hoi),
        ("Trùng lặp nội dung", trung_lap),
        ("Từ khóa tụt hạng", tut_hang),
        ("Từ khóa mới nổi", moi_noi),
        ("Trang hiệu quả nhất", trang_hieu_qua),
        ("Từ khóa hiệu quả nhất", tu_khoa_hieu_qua),
    ]:
        ws = wb.create_sheet(ten)
        _ghi_sheet(ws, df)

    ten_mien_sach = ten_mien.replace(".", "_").replace("https://", "").replace("http://", "")
    ten_file = f"{ten_mien_sach}_phan_tich_gsc_{ngay}.xlsx"
    duong_dan = thu_muc_xuat / ten_file
    thu_muc_xuat.mkdir(parents=True, exist_ok=True)
    wb.save(duong_dan)
    return duong_dan

# ---------------------------------------------------------------------------
# TẠO BÁO CÁO MARKDOWN
# ---------------------------------------------------------------------------
def tao_bao_cao_md(ten_mien, ngay, tong_quan, co_hoi, trung_lap, tut_hang, moi_noi):
    L = []
    L.append(f"# Báo cáo Phân tích GSC — {ten_mien}")
    L.append(f"\nNgày tạo: {ngay}\n")
    L.append("---\n")

    # I. TÓM TẮT ĐIỀU HÀNH
    L.append("## I. TÓM TẮT ĐIỀU HÀNH\n")
    L.append("| Chỉ số | Kỳ này | Kỳ trước | Thay đổi |")
    L.append("|--------|--------|----------|----------|")
    for chi_so, v in tong_quan.items():
        L.append(f"| {chi_so} | {v['Kỳ này']} | {v['Kỳ trước']} | {v['Thay đổi']} |")

    L.append("\n### Tín hiệu chính\n")
    L.append(f"- **{len(co_hoi)} cơ hội tối ưu CTR** — tăng clicks ngay không cần cải thiện thứ hạng")
    L.append(f"- **{len(trung_lap)} từ khóa trùng lặp nội dung** — đang phân tán lực SEO")
    L.append(f"- **{len(tut_hang)} từ khóa tụt hạng** — cần can thiệp")
    L.append(f"- **{len(moi_noi)} từ khóa/trang mới nổi** — cơ hội tăng trưởng 4–8 tuần tới\n")

    # II. PHÂN TÍCH WHAT–WHY–HOW
    L.append("---\n")
    L.append("## II. PHÂN TÍCH WHAT – WHY – HOW\n")

    # 2A
    L.append("### 2A. Cơ hội tối ưu CTR nhanh\n")
    if co_hoi.empty:
        L.append("_Không phát hiện cơ hội nào._\n")
    else:
        for _, r in co_hoi.head(5).iterrows():
            L.append(f"#### {r['Từ khóa']}")
            L.append(f"- **WHAT:** Vị trí {r['Vị trí TB']}, CTR {r['CTR thực tế']} vs chuẩn {r['CTR chuẩn']} — tiềm năng **+{r['Clicks tiềm năng']} clicks/kỳ**")
            L.append(f"- **WHY:** {r['Nguyên nhân CTR thấp']}")
            L.append(f"- **HOW:** {r['Hành động đề xuất']}\n")

    # 2B
    L.append("### 2B. Trùng lặp nội dung\n")
    if trung_lap.empty:
        L.append("_Không phát hiện trùng lặp._\n")
    else:
        for _, r in trung_lap.iterrows():
            L.append(f"#### {r['Từ khóa']} — {r['Mức độ nghiêm trọng']}")
            L.append(f"- **WHAT:** {r['Số URL cạnh tranh']} URL cạnh tranh — Trang chính: `{r['Trang chính (thứ hạng cao nhất)']}`")
            L.append(f"- **WHY:** {r['Nguyên nhân gây trùng lặp']}")
            L.append(f"- **HOW:** {r['Giải pháp đề xuất']}\n")

    # 2C
    L.append("### 2C. Từ khóa tụt hạng — Top 5\n")
    if tut_hang.empty:
        L.append("_Không phát hiện tụt hạng đáng kể._\n")
    else:
        for _, r in tut_hang.head(5).iterrows():
            L.append(f"#### {r['Từ khóa']} — {r['Mức độ nghiêm trọng']}")
            L.append(f"- **WHAT:** {r['Vị trí kỳ trước']} → {r['Vị trí kỳ này']} (tụt {r['Tụt (số bậc)']} bậc), thay đổi lượt nhấp: {r['Thay đổi lượt nhấp']}")
            L.append(f"- **WHY:** {r['Nguyên nhân có thể']}")
            L.append(f"- **HOW:** {r['Kế hoạch phục hồi']}\n")

    # 2D
    L.append("### 2D. Từ khóa mới nổi\n")
    if moi_noi.empty:
        L.append("_Chưa phát hiện từ khóa mới nổi đáng chú ý._\n")
    else:
        for _, r in moi_noi.head(5).iterrows():
            L.append(f"#### {r['Từ khóa']} — {r['Tiềm năng']}")
            L.append(f"- **WHAT:** {r['Loại']} — Vị trí {r['Vị trí kỳ này']}, lượt hiển thị {r['Lượt hiển thị']}, tăng {r['Tăng (số bậc)']} bậc")
            L.append(f"- **HOW:** {r['Hành động khai thác']}\n")

    # III. BẢNG CHI TIẾT
    L.append("---\n")
    L.append("## III. BẢNG CHI TIẾT\n")

    L.append(f"### 3A. Cơ hội tối ưu CTR ({len(co_hoi)} từ khóa)\n")
    if not co_hoi.empty:
        cols = ["Từ khóa", "Vị trí TB", "CTR thực tế", "CTR chuẩn", "Clicks tiềm năng"]
        L.append("| " + " | ".join(cols) + " |")
        L.append("|" + "|".join(["---"] * len(cols)) + "|")
        for _, r in co_hoi.head(10).iterrows():
            L.append("| " + " | ".join(str(r[c]) for c in cols) + " |")
        L.append("")

    L.append(f"### 3B. Trùng lặp nội dung ({len(trung_lap)} từ khóa)\n")
    if not trung_lap.empty:
        cols = ["Từ khóa", "Số URL cạnh tranh", "Vị trí trang chính", "Mức độ nghiêm trọng"]
        L.append("| " + " | ".join(cols) + " |")
        L.append("|" + "|".join(["---"] * len(cols)) + "|")
        for _, r in trung_lap.iterrows():
            L.append("| " + " | ".join(str(r[c]) for c in cols) + " |")
        L.append("")

    L.append(f"### 3C. Từ khóa tụt hạng ({len(tut_hang)} từ khóa)\n")
    if not tut_hang.empty:
        cols = ["Từ khóa", "Vị trí kỳ trước", "Vị trí kỳ này", "Tụt (số bậc)", "Mức độ nghiêm trọng"]
        L.append("| " + " | ".join(cols) + " |")
        L.append("|" + "|".join(["---"] * len(cols)) + "|")
        for _, r in tut_hang.head(10).iterrows():
            L.append("| " + " | ".join(str(r[c]) for c in cols) + " |")
        L.append("")

    L.append(f"### 3D. Từ khóa mới nổi ({len(moi_noi)} cơ hội)\n")
    if not moi_noi.empty:
        cols = ["Từ khóa", "Loại", "Vị trí kỳ này", "Tăng (số bậc)", "Lượt hiển thị", "Tiềm năng"]
        L.append("| " + " | ".join(cols) + " |")
        L.append("|" + "|".join(["---"] * len(cols)) + "|")
        for _, r in moi_noi.head(10).iterrows():
            L.append("| " + " | ".join(str(r[c]) for c in cols) + " |")
        L.append("")

    # IV. KẾ HOẠCH HÀNH ĐỘNG
    L.append("---\n")
    L.append("## IV. KẾ HOẠCH HÀNH ĐỘNG\n")
    L.append("| Độ ưu tiên | Hành động | Từ khóa/Trang | Kết quả kỳ vọng | Thời hạn |")
    L.append("|-----------|-----------|---------------|----------------|----------|")

    for _, r in trung_lap[trung_lap["Mức độ nghiêm trọng"] == "Nghiêm trọng"].head(2).iterrows():
        L.append(f"| P1 — Khẩn | Xử lý canonical/redirect | {r['Từ khóa']} | Tập trung lực SEO | Tuần này |")

    khung_hoang = tut_hang[tut_hang["Mức độ nghiêm trọng"].str.contains("Khủng hoảng", na=False)]
    for _, r in khung_hoang.head(2).iterrows():
        L.append(f"| P1 — Khẩn | Content audit + backlink | {r['Từ khóa']} | Phục hồi thứ hạng | 2 tuần |")

    for _, r in co_hoi.head(3).iterrows():
        L.append(f"| P2 — Quan trọng | Viết lại title/meta | {r['Từ khóa']} | +{r['Clicks tiềm năng']} clicks/kỳ | 1 tuần |")

    cao = moi_noi[moi_noi["Tiềm năng"].str.contains("Cao", na=False)]
    for _, r in cao.head(3).iterrows():
        L.append(f"| P3 — Cơ hội | Tối ưu nội dung + internal link | {r['Từ khóa']} | Tăng clicks hữu cơ | 4 tuần |")

    L.append("")
    L.append("---")
    L.append("_Báo cáo được tạo tự động bởi SEONGON GSC Insights Analyzer_")

    return "\n".join(L)

# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="SEONGON GSC Insights Analyzer")
    parser.add_argument("--domain", default="example.com")
    parser.add_argument("--mock", action="store_true")
    parser.add_argument("--output-dir", default=None)
    args = parser.parse_args()

    ngay = datetime.today().strftime("%Y%m%d")
    thu_muc_xuat = Path(args.output_dir) if args.output_dir else OUTPUTS_DIR
    thu_muc_xuat.mkdir(parents=True, exist_ok=True)

    if args.mock:
        print(f"[mock] Đang đọc dữ liệu mock: {args.domain}")
        df_hien_tai = doc_du_lieu_mock("mock_gsc_current.csv")
        df_truoc = doc_du_lieu_mock("mock_gsc_previous.csv")
    else:
        print("GSC API thực tế chưa triển khai. Dùng --mock.")
        sys.exit(1)

    print("Đang phân tích...")
    tong_quan   = tong_hop_toan_canh(df_hien_tai, df_truoc)
    co_hoi      = tim_co_hoi_toi_uu(df_hien_tai)
    trung_lap   = tim_trung_lap_noi_dung(df_hien_tai)
    tut_hang    = tim_tu_khoa_tut_hang(df_hien_tai, df_truoc)
    moi_noi     = tim_tu_khoa_moi_noi(df_hien_tai, df_truoc)
    trang_hq    = phan_tich_trang_hieu_qua(df_hien_tai)
    tu_khoa_hq  = phan_tich_tu_khoa_hieu_qua(df_hien_tai)

    print(f"  Cơ hội tối ưu CTR:  {len(co_hoi)}")
    print(f"  Trùng lặp nội dung: {len(trung_lap)}")
    print(f"  Từ khóa tụt hạng:   {len(tut_hang)}")
    print(f"  Từ khóa mới nổi:    {len(moi_noi)}")

    duong_dan_excel = xuat_excel(
        ten_mien=args.domain, tong_quan=tong_quan, co_hoi=co_hoi,
        trung_lap=trung_lap, tut_hang=tut_hang, moi_noi=moi_noi,
        trang_hieu_qua=trang_hq, tu_khoa_hieu_qua=tu_khoa_hq,
        thu_muc_xuat=thu_muc_xuat, ngay=ngay,
    )
    print(f"\nXuất Excel: {duong_dan_excel}")

    noi_dung_md = tao_bao_cao_md(
        ten_mien=args.domain, ngay=ngay, tong_quan=tong_quan,
        co_hoi=co_hoi, trung_lap=trung_lap, tut_hang=tut_hang, moi_noi=moi_noi,
    )
    ten_mien_sach = args.domain.replace(".", "_")
    duong_dan_md = thu_muc_xuat / f"{ten_mien_sach}_bao_cao_gsc_{ngay}.md"
    duong_dan_md.write_text(noi_dung_md, encoding="utf-8")
    print(f"Xuất Markdown: {duong_dan_md}")
    print("\nHoàn tất!")


if __name__ == "__main__":
    main()
